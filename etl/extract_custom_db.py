#!/usr/bin/env python3
"""Extract and normalize the Functional Fitness Exercise Database (v2.9) spreadsheet.

Reads the .xlsx, normalizes the 31-column taxonomy, resolves the YouTube
hyperlinks behind the "Video Demonstration" / "Video Explanation" cells, maps a
subset onto wger's taxonomy, and writes:

  build/exercises.jsonl   one normalized exercise per line (sidecar source of truth)
  build/qc_report.json    every data-quality issue found, grouped by kind
  build/qc_report.md      the same report, human-readable

The spreadsheet's own layout facts (header on row 16, data in columns B..AF) are
asserted rather than assumed, so a future v3.0 with a different shape fails loudly
instead of silently producing garbage.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
import uuid
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

# The spreadsheet has 15 rows of preamble (download links, YouTube tutorial,
# changelog) before the real header row.
HEADER_ROW = 16
FIRST_DATA_ROW = 17

# Column letter -> field name. Trailing spaces in several spreadsheet headers
# ("Target Muscle Group ") are why we key on position and assert on the text.
COLUMNS = {
    "B": ("name", "Exercise"),
    "C": ("video_demo_url", "Short YouTube Demonstration"),
    "D": ("video_explain_url", "In-Depth YouTube Explanation"),
    "E": ("difficulty", "Difficulty Level"),
    "F": ("target_muscle_group", "Target Muscle Group"),
    "G": ("prime_mover_muscle", "Prime Mover Muscle"),
    "H": ("secondary_muscle", "Secondary Muscle"),
    "I": ("tertiary_muscle", "Tertiary Muscle"),
    "J": ("primary_equipment", "Primary Equipment"),
    "K": ("primary_items", "# Primary Items"),
    "L": ("secondary_equipment", "Secondary Equipment"),
    "M": ("secondary_items", "# Secondary Items"),
    "N": ("posture", "Posture"),
    "O": ("arm_involvement", "Single or Double Arm"),
    "P": ("arm_action", "Continuous or Alternating Arms"),
    "Q": ("grip", "Grip"),
    "R": ("load_position", "Load Position (Ending)"),
    "S": ("leg_action", "Continuous or Alternating Legs"),
    "T": ("foot_elevation", "Foot Elevation"),
    "U": ("combination", "Combination Exercises"),
    "V": ("movement_pattern_1", "Movement Pattern #1"),
    "W": ("movement_pattern_2", "Movement Pattern #2"),
    "X": ("movement_pattern_3", "Movement Pattern #3"),
    "Y": ("plane_1", "Plane Of Motion #1"),
    "Z": ("plane_2", "Plane Of Motion #2"),
    "AA": ("plane_3", "Plane Of Motion #3"),
    "AB": ("body_region", "Body Region"),
    "AC": ("force_type", "Force Type"),
    "AD": ("mechanics", "Mechanics"),
    "AE": ("laterality", "Laterality"),
    "AF": ("classification", "Primary Exercise Classification"),
}

# Values that mean "the author has not categorized this yet", not a real value.
PLACEHOLDERS = {"", "unsorted*", "n/a", "na", "-", "none specified"}

# Enum whitelists for the columns where the spreadsheet contains provably invalid
# entries. Anything outside these is nulled and flagged rather than propagated.
VALID_MECHANICS = {"Compound", "Isolation"}
VALID_PLANES = {"Sagittal Plane", "Frontal Plane", "Transverse Plane"}

# Values that look like data corruption but are not.
#
# `Load Position (Ending)` = "Order" on 192 rows initially looked like a paste
# artifact. It isn't: "order position" is a real clubbell/Indian-club position (the
# club held vertically, from the military drill "order arms"), which is why it appears
# in exercise names too — "Double Clubbell Order Squat", "Clubbell Order Cyclist
# Squat". Confirmed as legitimate by the database owner, so it is neither nulled nor
# flagged. Left here as documentation so it doesn't get "fixed" by a future reader.
SUSPICIOUS_VALUES: dict[str, set[str]] = {}

# Stable namespace for deterministic UUIDv5 generation. Exercises imported into
# wger get a UUID derived from this namespace + slug, so re-running the import
# updates rather than duplicates, and no generated UUID can ever collide with an
# upstream wger.de exercise UUID.
UUID_NAMESPACE = uuid.uuid5(uuid.NAMESPACE_URL, "https://github.com/ai-fitness-selfhost/custom-exercise-db")

MAPPINGS_PATH = Path(__file__).parent / "mappings" / "wger_mappings.json"


def clean(value) -> str | None:
    """Normalize a raw cell into a trimmed string, or None if it carries no value."""
    if value is None:
        return None
    text = str(value)
    # NBSP and zero-width characters appear in several cells and would otherwise
    # survive .strip() and break exact-match lookups.
    text = text.replace(" ", " ").replace("​", "")
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"\s+", " ", text).strip()
    if text.lower() in PLACEHOLDERS:
        return None
    return text or None


def slugify(name: str) -> str:
    slug = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", slug).strip("-").lower()
    return slug


def as_int(value) -> int | None:
    text = clean(value)
    if text is None:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def verify_layout(ws) -> list[str]:
    """Assert the spreadsheet still has the shape this script was written for."""
    problems = []
    for letter, (field, expected) in COLUMNS.items():
        actual = clean(ws[f"{letter}{HEADER_ROW}"].value)
        if actual is None or actual.lower() != expected.lower():
            problems.append(
                f"column {letter}: expected header {expected!r}, found {actual!r} "
                f"(field {field})"
            )
    return problems


def extract(xlsx_path: Path) -> tuple[list[dict], dict]:
    # data_only=True resolves formulas to values; hyperlinks require the
    # non-read-only loader, which is why we accept the slower full load.
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Exercises" not in workbook.sheetnames:
        sys.exit(f"error: no 'Exercises' sheet in {xlsx_path.name} (found {workbook.sheetnames})")
    ws = workbook["Exercises"]

    layout_problems = verify_layout(ws)
    if layout_problems:
        sys.exit(
            "error: spreadsheet layout does not match this extractor.\n  "
            + "\n  ".join(layout_problems)
        )

    mappings = json.loads(MAPPINGS_PATH.read_text())
    category_map = mappings["category"]
    muscle_map = mappings["muscle"]
    equipment_existing = mappings["equipment_existing"]
    difficulty_rank = mappings["difficulty_rank"]

    exercises: list[dict] = []
    issues: dict[str, list] = defaultdict(list)
    name_counts: Counter[str] = Counter()
    slug_counts: Counter[str] = Counter()

    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        raw_name = ws[f"B{row}"].value
        if raw_name is None or not str(raw_name).strip():
            continue

        name = clean(raw_name)
        if name is None:
            issues["unusable_name"].append({"row": row, "raw": repr(raw_name)})
            continue

        if str(raw_name) != name:
            issues["name_whitespace"].append({"row": row, "raw": repr(str(raw_name)), "cleaned": name})

        record: dict = {"source": "ffed-2.9", "source_row": row, "name": name}
        qc_flags: list[str] = []

        for letter, (field, _) in COLUMNS.items():
            if field in ("name", "video_demo_url", "video_explain_url"):
                continue
            record[field] = clean(ws[f"{letter}{row}"].value)

        # --- video hyperlinks -------------------------------------------------
        # The cells display the words "Video Demonstration" / "Video
        # Explanation"; the real YouTube URL is the cell's hyperlink target.
        for letter, field in (("C", "video_demo_url"), ("D", "video_explain_url")):
            cell = ws[f"{letter}{row}"]
            link = cell.hyperlink.target if cell.hyperlink else None
            if link is None and clean(cell.value):
                # Label text present but no hyperlink attached — the link was lost.
                qc_flags.append(f"{field}_label_without_link")
            record[field] = link.strip() if link else None

        # --- counts -----------------------------------------------------------
        record["primary_items"] = as_int(ws[f"K{row}"].value)
        record["secondary_items"] = as_int(ws[f"M{row}"].value)

        # --- enum validation --------------------------------------------------
        if record["mechanics"] is not None and record["mechanics"] not in VALID_MECHANICS:
            issues["invalid_mechanics"].append(
                {"row": row, "name": name, "value": record["mechanics"]}
            )
            qc_flags.append("invalid_mechanics")
            record["mechanics"] = None
        elif record["mechanics"] is None:
            qc_flags.append("missing_mechanics")

        for field, bad_values in SUSPICIOUS_VALUES.items():
            if record.get(field) in bad_values:
                issues["suspicious_value"].append(
                    {"row": row, "name": name, "field": field, "value": record[field]}
                )
                qc_flags.append(f"suspicious_{field}")

        # --- collapse the numbered multi-value columns -------------------------
        patterns = []
        for key in ("movement_pattern_1", "movement_pattern_2", "movement_pattern_3"):
            value = record.pop(key)
            # "Other" carries no information in this column and is dropped.
            if value and value != "Other" and value not in patterns:
                patterns.append(value)
        record["movement_patterns"] = patterns
        if not patterns:
            qc_flags.append("no_movement_pattern")

        planes = []
        for key in ("plane_1", "plane_2", "plane_3"):
            value = record.pop(key)
            if value is None:
                continue
            if value not in VALID_PLANES:
                issues["invalid_plane"].append({"row": row, "name": name, "value": value})
                qc_flags.append("invalid_plane")
                continue
            if value not in planes:
                planes.append(value)
        record["planes_of_motion"] = planes
        if not planes:
            qc_flags.append("no_plane_of_motion")

        # --- derived fields ---------------------------------------------------
        record["is_combo"] = record.pop("combination") == "Combo Exercise"
        record["difficulty_rank"] = difficulty_rank.get(record["difficulty"] or "")
        if record["difficulty_rank"] is None:
            issues["unknown_difficulty"].append(
                {"row": row, "name": name, "value": record["difficulty"]}
            )
            qc_flags.append("unknown_difficulty")

        # Placeholder-only columns get flagged so the AI backfill can prioritize them.
        for field in ("force_type", "classification"):
            if record[field] is None:
                qc_flags.append(f"unsorted_{field}")

        # --- wger mapping -----------------------------------------------------
        wger_category = category_map.get(record["target_muscle_group"] or "")
        if wger_category is None:
            issues["unmapped_category"].append(
                {"row": row, "name": name, "value": record["target_muscle_group"]}
            )
            qc_flags.append("unmapped_category")

        primary_wger_muscles: list[int] = []
        secondary_wger_muscles: list[int] = []
        for field, bucket in (
            ("prime_mover_muscle", primary_wger_muscles),
            ("secondary_muscle", secondary_wger_muscles),
            ("tertiary_muscle", secondary_wger_muscles),
        ):
            muscle = record[field]
            if muscle is None:
                continue
            if muscle not in muscle_map:
                issues["unmapped_muscle"].append({"row": row, "name": name, "value": muscle})
                qc_flags.append("unmapped_muscle")
                continue
            mapped = muscle_map[muscle]
            # null is a deliberate "no wger equivalent", not a mapping gap.
            if mapped is not None and mapped not in bucket:
                bucket.append(mapped)
        # A muscle that is primary must not also be listed as secondary in wger.
        secondary_wger_muscles = [m for m in secondary_wger_muscles if m not in primary_wger_muscles]

        wger_equipment: list[int] = []
        unmapped_equipment: list[str] = []
        for field in ("primary_equipment", "secondary_equipment"):
            item = record[field]
            # "None" in Secondary Equipment means genuinely no second implement.
            if item is None or item == "None":
                continue
            if item in equipment_existing:
                if equipment_existing[item] not in wger_equipment:
                    wger_equipment.append(equipment_existing[item])
            else:
                unmapped_equipment.append(item)

        record["wger"] = {
            "category": wger_category,
            "muscles": sorted(primary_wger_muscles),
            "muscles_secondary": sorted(secondary_wger_muscles),
            "equipment": sorted(wger_equipment),
            # Resolved to real wger equipment IDs by the import command, after it
            # creates the local equipment rows listed in the mappings file.
            "equipment_pending_creation": sorted(set(unmapped_equipment)),
        }

        # --- identity ---------------------------------------------------------
        slug = slugify(name)
        slug_counts[slug] += 1
        name_counts[name] += 1
        if slug_counts[slug] > 1:
            # Deterministic disambiguation so re-runs are stable.
            slug = f"{slug}-{slug_counts[slug]}"
            qc_flags.append("duplicate_name")
        record["slug"] = slug
        record["uuid"] = str(uuid.uuid5(UUID_NAMESPACE, slug))

        # Populated later by the AI description backfill; wger requires >= 40 chars.
        record["description"] = None
        record["qc_flags"] = qc_flags
        exercises.append(record)

    for name, count in name_counts.items():
        if count > 1:
            issues["duplicate_name"].append({"name": name, "occurrences": count})

    stats = {
        "source_file": xlsx_path.name,
        "exercises": len(exercises),
        "with_demo_video": sum(1 for e in exercises if e["video_demo_url"]),
        "with_explainer_video": sum(1 for e in exercises if e["video_explain_url"]),
        "with_description": sum(1 for e in exercises if e["description"]),
        "clean_records": sum(1 for e in exercises if not e["qc_flags"]),
        "flagged_records": sum(1 for e in exercises if e["qc_flags"]),
        "flag_counts": dict(Counter(f for e in exercises for f in e["qc_flags"]).most_common()),
        "equipment_needing_wger_rows": sorted(
            {i for e in exercises for i in e["wger"]["equipment_pending_creation"]}
        ),
        "issue_counts": {kind: len(rows) for kind, rows in sorted(issues.items())},
    }
    return exercises, {"stats": stats, "issues": {k: v for k, v in sorted(issues.items())}}


def write_markdown_report(report: dict, path: Path) -> None:
    stats = report["stats"]
    lines = [
        "# Exercise database extraction report",
        "",
        f"Source: `{stats['source_file']}`",
        "",
        "## Coverage",
        "",
        "| Metric | Count |",
        "|---|---|",
        f"| Exercises extracted | {stats['exercises']} |",
        f"| With demonstration video | {stats['with_demo_video']} |",
        f"| With explainer video | {stats['with_explainer_video']} |",
        f"| With written description | {stats['with_description']} |",
        f"| Records with no QC flags | {stats['clean_records']} |",
        f"| Records with >= 1 QC flag | {stats['flagged_records']} |",
        "",
        "## QC flags by frequency",
        "",
        "| Flag | Records |",
        "|---|---|",
    ]
    lines += [f"| `{flag}` | {count} |" for flag, count in stats["flag_counts"].items()]
    lines += [
        "",
        "## Equipment requiring new wger rows",
        "",
        f"{len(stats['equipment_needing_wger_rows'])} equipment types in this database have no "
        "wger equivalent. wger's `/api/v2/equipment/` endpoint is read-only, so these are created "
        "by the import management command rather than over the REST API.",
        "",
    ]
    lines += [f"- {name}" for name in stats["equipment_needing_wger_rows"]]
    lines += ["", "## Issues requiring a human decision", ""]

    if not report["issues"]:
        lines.append("None.")
    for kind, rows in report["issues"].items():
        lines += [f"### `{kind}` ({len(rows)})", ""]
        for row in rows[:15]:
            lines.append(f"- {json.dumps(row, ensure_ascii=False)}")
        if len(rows) > 15:
            lines.append(f"- _... and {len(rows) - 15} more (see `qc_report.json`)_")
        lines.append("")

    path.write_text("\n".join(lines) + "\n")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "xlsx",
        nargs="?",
        default="Functional+Fitness+Exercise+Database+(version+2.9).xlsx",
        help="path to the exercise database spreadsheet",
    )
    parser.add_argument("--out", default="build", help="output directory (default: build)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"error: {xlsx_path} not found")

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    exercises, report = extract(xlsx_path)

    jsonl_path = out_dir / "exercises.jsonl"
    with jsonl_path.open("w") as handle:
        for exercise in exercises:
            handle.write(json.dumps(exercise, ensure_ascii=False) + "\n")

    (out_dir / "qc_report.json").write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n")
    write_markdown_report(report, out_dir / "qc_report.md")

    stats = report["stats"]
    print(f"extracted {stats['exercises']} exercises -> {jsonl_path}")
    print(f"  {stats['clean_records']} clean, {stats['flagged_records']} flagged")
    print(f"  {stats['with_demo_video']} demo videos, {stats['with_explainer_video']} explainers")
    print(f"  {len(stats['equipment_needing_wger_rows'])} equipment types need new wger rows")
    print(f"report -> {out_dir / 'qc_report.md'}")


if __name__ == "__main__":
    main()
